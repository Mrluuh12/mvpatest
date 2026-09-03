"""Leitura da planilha de inventário.

A planilha real tem uma linha de título antes do cabeçalho, e o cabeçalho não
está sempre na mesma linha entre abas. Em vez de fixar índices — que quebram no
dia em que alguém insere uma linha —, o cabeçalho é localizado pelo conteúdo.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl

#: Colunas que precisam existir para a linha ser considerada um cabeçalho.
_MARCADORES = ("name", "ip")

#: Nome da aba consolidada. Se não existir, cai na primeira aba do arquivo.
ABA_PADRAO = "Inventário Completo"


def _e_cabecalho(linha: tuple[Any, ...]) -> bool:
    celulas = {str(c).strip().lower() for c in linha if c is not None}
    return all(m in celulas for m in _MARCADORES)


def ler(caminho: str | Path, aba: str | None = None) -> list[dict[str, Any]]:
    """Lê a planilha e devolve os registros como dicionários.

    Levanta ``ValueError`` se nenhum cabeçalho for encontrado — falhar aqui é
    melhor que devolver uma lista vazia que parece "nenhum ativo".
    """
    livro = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    nome_aba = aba or (ABA_PADRAO if ABA_PADRAO in livro.sheetnames else livro.sheetnames[0])
    folha = livro[nome_aba]

    linhas = list(folha.iter_rows(values_only=True))
    indice_cabecalho = next((i for i, ln in enumerate(linhas[:10]) if _e_cabecalho(ln)), None)
    if indice_cabecalho is None:
        raise ValueError(
            f"cabeçalho não encontrado na aba {nome_aba!r}: "
            f"esperava colunas {_MARCADORES}"
        )

    cabecalho = [
        str(c).strip() if c is not None else f"coluna_{i}"
        for i, c in enumerate(linhas[indice_cabecalho])
    ]

    registros: list[dict[str, Any]] = []
    for linha in linhas[indice_cabecalho + 1 :]:
        if not any(c is not None for c in linha):
            continue
        registro = {
            chave: (str(valor).strip() if valor is not None else None)
            for chave, valor in zip(cabecalho, linha, strict=False)
        }
        if registro.get("Name"):
            registros.append(registro)
    return registros
